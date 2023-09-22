from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Student_data, Faculty_profile, HOD_profile, UploadedFiles, StudentFacultyMapping
from django.http import HttpResponse
import openpyxl

def main_dashboard(request):
    return render(request, "main_dashboard.html")

# def Student_Login(request):
#     if request.method == "POST":
#         username = request.POST.get("username")
#         password = request.POST.get("password")
#
#         try:
#             student = Student_data.objects.get(username=username, password=password)
#             return redirect('student_dashboard')
#         except Student_data.DoesNotExist:
#             error_message = "Invalid username or password"
#             return render(request, "Student_Login.html", {"error_message": error_message})
#
#     return render(request, "Student_Login.html")
def Student_Login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        try:
            # Check if the student exists in the Student_data model
            student = Student_data.objects.get(username=username, password=password)

            # Check if there is a mapping for this student in StudentFacultyMapping
            try:
                mapping = StudentFacultyMapping.objects.get(student_name=student.username)

                # If a mapping exists, retrieve faculty information
                faculty_name = mapping.faculty_name
                faculty_email = mapping.faculty_email
                faculty_phone = mapping.faculty_phone

                # Render the student dashboard with faculty information
                return render(request, "student_dashboard.html", {
                    "username": username,
                    "faculty_name": faculty_name,
                    "faculty_email": faculty_email,
                    "faculty_phone": faculty_phone,
                })
            except StudentFacultyMapping.DoesNotExist:
                # Handle the case where there is no mapping for this student
                error_message = "Student has no faculty mapping"
                return render(request, "Student_Login.html", {"error_message": error_message})

        except Student_data.DoesNotExist:
            error_message = "Invalid username or password"
            return render(request, "Student_Login.html", {"error_message": error_message})

    return render(request, "Student_Login.html")


def student_dashboard(request):
    # Redirect to login page if not authenticated
    if not request.user.is_authenticated:
        return render(request, "Student_Login.html")

    # If authenticated, display the student dashboard
    return render(request, "student_dashboard.html")

def student_register(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        username = request.POST.get("username")
        password = request.POST.get("password")

        # Check if the username already exists in the database
        existing_student = Student_data.objects.filter(username=username).first()

        if existing_student:
            # If a student with the same username already exists, handle the error here
            error_message = "Username already exists. Please choose a different username."
            return render(request, "Student_Login.html", {"error_message": error_message})

        # If the username is unique, create a new student record
        student = Student_data(
            first_name=first_name,
            last_name=last_name,
            username=username,
            password=password
        )
        student.save()

        return redirect('Student_Login')

    return render(request, "Student_Login.html")


# def faculty_login(request):
#     if request.method == "POST":
#         username = request.POST.get("username")
#         password = request.POST.get("password")
#
#         try:
#             faculty = Faculty_profile.objects.get(username=username, password=password)
#             # If the login is successful, redirect to the faculty_dashboard URL
#             return redirect('faculty_dashboard',{"username": username})
#         except Faculty_profile.DoesNotExist:
#             error_message = "Invalid username or password"
#             return render(request, "Faculty_Login.html", {"error_message": error_message})
#
#     return render(request, "Faculty_Login.html")

def faculty_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        try:
            faculty = Faculty_profile.objects.get(username=username, password=password)
            # If the login is successful, store the username in the session
            request.session['username'] = username
            return redirect('faculty_dashboard')
        except Faculty_profile.DoesNotExist:
            error_message = "Invalid username or password"
            #messages.error(request, error_message)
            return render(request, "Faculty_Login.html")

    return render(request, "Faculty_Login.html")


def faculty_dashboard(request):
    username = request.session.get('username')  # Retrieve the username from the session

    # Filter faculty_mappings based on the logged-in username
    faculty_mappings = StudentFacultyMapping.objects.filter(faculty_name=username)

    return render(request, 'faculty_dashboard.html', {'faculty_mappings': faculty_mappings, 'username': username})
# def faculty_dashboard(request):
#     username = request.user.username
#     student_faculty_mappings = StudentFacultyMapping.objects.filter(faculty_name=username)
#
#     context = {
#         'username': username,
#         'student_faculty_mappings': student_faculty_mappings,
#     }
#
#     return render(request, 'faculty_dashboard.html', context)


def hod_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        try:
            hod = HOD_profile.objects.get(username=username, password=password)
            return render(request, 'hod_dashboard.html')
        except HOD_profile.DoesNotExist:
            error_message = "Invalid username or password"
            return render(request, "HOD_Login.html", {"error_message": error_message})

    return render(request, "HOD_Login.html")

@login_required
def hod_dashboard(request):
    if request.method == 'POST':
        excel_file1 = request.FILES.get('excel_file1')
        excel_file2 = request.FILES.get('excel_file2')

        if excel_file1 and excel_file2:
            uploaded_files = UploadedFiles(file1=excel_file1, file2=excel_file2)
            uploaded_files.save()

    return render(request, 'hod_dashboard.html')


# def upload_files(request):
#     if request.method == 'POST':
#         excel_file1 = request.FILES.get('excel_file1')
#         excel_file2 = request.FILES.get('excel_file2')
#
#         if excel_file1 and excel_file2:
#             uploaded_files = UploadedFiles(file1=excel_file1, file2=excel_file2)
#             uploaded_files.save()
#
#     return render(request, 'hod_dashboard.html')

def generate_mapping_excel(request):
    if request.method == 'POST':
        faculty_file = request.FILES.get('faculty_file')
        student_file = request.FILES.get('student_file')

        if faculty_file and student_file:
            faculty_wb = openpyxl.load_workbook(faculty_file)
            faculty_sheet = faculty_wb.active

            student_wb = openpyxl.load_workbook(student_file)
            student_sheet = student_wb.active

            mapping = {'Professor': 15, 'Associate Professor': 15, 'Asst.Prof': 10, 'HOD': 15}
            faculty_mapping = {}

            faculty_rows = list(faculty_sheet.iter_rows(values_only=True))
            student_rows = list(student_sheet.iter_rows(values_only=True))

            faculty_rows.pop(0)  # Remove header row
            student_rows.pop(0)  # Remove header row

            for faculty_row in faculty_rows:
                designation = faculty_row[4]
                if designation in mapping:
                    num_students = mapping[designation]
                    assigned_students = student_rows[:num_students]
                    student_rows = student_rows[num_students:]  # Update student_rows for the next faculty member
                    faculty_mapping[faculty_row[1]] = {
                        'faculty_info': faculty_row,
                        'assigned_students': assigned_students
                    }

            mapped_wb = openpyxl.Workbook()
            mapped_sheet = mapped_wb.active
            mapped_sheet.append(
                ['S.No', 'Registration Number', 'Name', 'Branch/Specialization', 'Student Email', 'Faculty Name',
                 'Faculty Email', 'Faculty Phone'])

            s_no = 1

            for faculty_name, faculty_data in faculty_mapping.items():
                faculty_info = faculty_data['faculty_info']
                assigned_students = faculty_data['assigned_students']

                for student_row in assigned_students:
                    student_email = student_row[6]  # Assuming the student email is in the 7th column

                    # Check if student_email is not empty before creating the mapping entry
                    if student_email:
                        faculty_email = faculty_info[6]
                        faculty_phone = faculty_info[5]

                        mapped_sheet.append([
                            s_no,
                            student_row[1],
                            student_row[2],
                            student_row[3],
                            student_row[6],
                            faculty_name,
                            faculty_email,
                            faculty_phone
                        ])
                        s_no += 1

                        # Create a mapping entry and save to the database
                        mapping_entry = StudentFacultyMapping(
                            student_registration_number=student_row[1],
                            student_name=student_row[2],
                            student_branch=student_row[3],
                            student_email=student_email,
                            faculty_name=faculty_name,
                            faculty_email=faculty_email,
                            faculty_phone=faculty_phone
                        )
                        mapping_entry.save()

            excel_response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_response['Content-Disposition'] = 'attachment; filename=student_faculty_mapping.xlsx'
            mapped_wb.save(excel_response)

            return excel_response

    return render(request, 'hod_dashboard.html')


def download_mapped_excel(request):
    # Retrieve data from the StudentFacultyMapping model
    mapping_data = StudentFacultyMapping.objects.all()

    # Prepare Excel data
    mapped_wb = openpyxl.Workbook()
    mapped_sheet = mapped_wb.active
    mapped_sheet.append(
        ['S.No', 'Registration Number', 'Name', 'Branch/Specialization', 'Student Email', 'Faculty Name',
         'Faculty Email', 'Faculty Phone'])

    for index, entry in enumerate(mapping_data, start=1):
        mapped_sheet.append([
            index,
            entry.student_registration_number,
            entry.student_name,
            entry.student_branch,
            entry.student_email,
            entry.faculty_name,
            entry.faculty_email,
            entry.faculty_phone
        ])

    excel_response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    excel_response['Content-Disposition'] = 'attachment; filename=student_faculty_mapping.xlsx'
    mapped_wb.save(excel_response)

    return excel_response

def mapping_list(request):
    mappings = StudentFacultyMapping.objects.all()
    return render(request, 'mapping_list.html', {'mappings': mappings})

from django.contrib.auth.decorators import login_required

@login_required  # Apply the login_required decorator
def student_files(request):
    if request.method == 'POST':
        weekly_report = request.FILES.get('weekly_report')
        certificate = request.FILES.get('certificate')
        ppt = request.FILES.get('ppt')
        final_project_report = request.FILES.get('final_project_report')

        if weekly_report and certificate and ppt and final_project_report:
            # Create a new entry in the MyUploadedFiles model and associate it with the logged-in student
            uploaded_files = MyUploadedFiles(
                weekly_report=weekly_report,
                certificate=certificate,
                ppt=ppt,
                final_project_report=final_project_report,
                student=request.user,  # Associate the file with the logged-in student
            )
            uploaded_files.save()
            return redirect('student_files')  # Redirect back to the same page

    return render(request, 'student_files.html')


from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404
from .models import MyUploadedFiles
import os


def download_file(request, file_id):
    # Use get_object_or_404 to handle file not found
    uploaded_file = get_object_or_404(MyUploadedFiles, id=file_id)

    file_path = uploaded_file.weekly_report.path  # Assuming this is the correct path

    # Check if the file exists
    if os.path.exists(file_path):
        # Open and serve the file
        with open(file_path, 'rb') as file:
            response = FileResponse(file)
        return response
    else:
        raise Http404("File not found")

